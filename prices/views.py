import csv
import os
import re
from datetime import datetime
from openpyxl import load_workbook

import pdfplumber

from rest_framework import generics, permissions, status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from .models import PriceUpload, CropPrice
from .serializers import PriceUploadSerializer


class IsAdminOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)


class PriceUploadListCreateAPI(generics.ListCreateAPIView):
    queryset = PriceUpload.objects.all().order_by("-created_at")
    serializer_class = PriceUploadSerializer
    permission_classes = [IsAdminOnly]
    parser_classes = [MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

        upload = PriceUpload.objects.create(
            file=file,
            original_name=file.name,
            uploaded_by=request.user,
            status="PENDING",
        )

        try:
            rows = self._parse_and_save(upload)
            upload.status = "PROCESSED"
            upload.processed_rows = rows
            upload.save()
        except Exception as e:
            upload.status = "FAILED"
            upload.error_message = str(e)
            upload.save()

        serializer = self.get_serializer(upload, context={"request": request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def _parse_and_save(self, upload: PriceUpload) -> int:
        """
        Expected columns (recommended):
        crop_name, date, price, unit, market

        date format: YYYY-MM-DD
        """
        path = upload.file.path
        ext = os.path.splitext(path)[1].lower()

        data_rows = []

        if ext == ".csv":
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    data_rows.append(r)

        elif ext in [".xlsx", ".xlsm"]:
            wb = load_workbook(path)
            ws = wb.active
            headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                r = {headers[i]: row[i] for i in range(len(headers))}
                data_rows.append(r)

        elif ext == ".pdf":
            data_rows.extend(self._extract_pdf_rows(path))

        else:
            raise Exception("Only CSV, Excel, and PDF files are supported")

        created = 0
        for r in data_rows:
            crop_name = (r.get("crop_name") or r.get("Crop Name") or r.get("crop") or "").strip()
            date_raw = r.get("date") or r.get("Date")
            price_raw = r.get("price") or r.get("Price")
            unit = r.get("unit") or r.get("Unit")
            market = r.get("market") or r.get("Market")

            if not crop_name or not date_raw or price_raw in [None, ""]:
                continue

            # date parsing
            date_val = self._parse_date(date_raw)
            price_val = self._parse_price(price_raw)

            CropPrice.objects.create(
                crop_name=crop_name,
                date=date_val,
                price=price_val,
                unit=str(unit).strip() if unit else None,
                market=str(market).strip() if market else None,
                upload=upload,
            )
            created += 1

        return created

    def _extract_pdf_rows(self, path: str) -> list[dict]:
        data_rows = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    table_rows = [row for row in table if row and any((cell or "").strip() for cell in row)]
                    if len(table_rows) < 2:
                        continue

                    headers = [self._normalize_header(cell) for cell in table_rows[0]]
                    if not {"crop_name", "date", "price"}.issubset(set(headers)):
                        continue

                    for row in table_rows[1:]:
                        values = list(row) + [None] * (len(headers) - len(row))
                        mapped = {headers[i]: (values[i] or "").strip() if isinstance(values[i], str) else values[i] for i in range(len(headers))}
                        data_rows.append(mapped)

                # Fallback for simple line-based PDFs (non-table extraction)
                if not tables:
                    text = page.extract_text() or ""
                    for line in text.splitlines():
                        parsed = self._parse_pdf_text_line(line)
                        if parsed:
                            data_rows.append(parsed)

        return data_rows

    def _parse_pdf_text_line(self, line: str) -> dict | None:
        # Accept lines such as:
        # Tomato,2026-03-12,180.50,kg,Dambulla
        # Tomato 2026-03-12 180.50 kg Dambulla
        line = (line or "").strip()
        if not line:
            return None

        date_match = re.search(r"\b\d{4}-\d{2}-\d{2}\b", line)
        price_match = re.search(r"(?<!\d)(\d+(?:\.\d{1,2})?)(?!\d)", line)
        if not date_match or not price_match:
            return None

        date_raw = date_match.group(0)
        price_raw = price_match.group(1)

        # Crop name assumed to be before the date token.
        crop_name = line[: date_match.start()].strip(" ,-|\t")
        if not crop_name:
            return None

        tail = line[date_match.end():].strip(" ,-|\t")
        unit = None
        market = None

        if tail:
            tokens = [t for t in re.split(r"[\s,]+", tail) if t]
            if len(tokens) >= 2:
                unit = tokens[1] if tokens[0] == price_raw else tokens[0]
            if len(tokens) >= 3:
                market = " ".join(tokens[2:]) if tokens[0] == price_raw else " ".join(tokens[1:])

        return {
            "crop_name": crop_name,
            "date": date_raw,
            "price": price_raw,
            "unit": unit,
            "market": market,
        }

    def _normalize_header(self, header):
        value = (str(header or "").strip().lower())
        key = value.replace(" ", "_")

        aliases = {
            "crop": "crop_name",
            "crop_name": "crop_name",
            "date": "date",
            "price": "price",
            "unit": "unit",
            "market": "market",
        }
        return aliases.get(key, key)

    def _parse_date(self, raw_value):
        if isinstance(raw_value, datetime):
            return raw_value.date()

        text = str(raw_value).strip()
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"]:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue

        raise ValueError(f"Unsupported date format: {text}")

    def _parse_price(self, raw_value):
        text = str(raw_value).strip().replace(",", "")
        return float(text)
